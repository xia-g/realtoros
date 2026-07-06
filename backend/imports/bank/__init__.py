"""Bank Import — parsers for CSV, XLSX, 1C export, MT940, CAMT.053.

ImportResult.import_file(file) → batch_id, events, duplicates, warnings, preview.

No core table changes — only adapters.
"""

from __future__ import annotations

import csv
import hashlib
import io
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, BinaryIO

from backend.accounting.db.pool import get_pool
from backend.accounting.models.enums import EventType


@dataclass
class BankRow:
    """Normalized bank transaction row."""
    external_id: str | None
    amount: float
    currency: str
    direction: str  # inflow | outflow
    transaction_date: date
    value_date: date | None
    counterparty: str | None
    account: str | None
    description: str | None
    purpose: str | None
    reference: str | None
    source_format: str  # csv, xlsx, mt940, camt053, 1c
    raw: dict | None = None


@dataclass
class ImportResult:
    batch_id: str
    events_created: int
    duplicates: int
    warnings: list[str]
    preview: list[BankRow]
    fingerprint: str | None = None


# ── Parser Registry ────────────────────────────────────────────────────

PARSERS: dict[str, str] = {
    "csv": "parse_csv",
    "xlsx": "parse_xlsx",
    "1c_export": "parse_1c",
    "mt940": "parse_mt940",
    "camt053": "parse_camt053",
}


def detect_format(filename: str, content: bytes | None = None) -> str:
    """Detect bank file format from extension and content."""
    ext = filename.lower().split(".")[-1] if "." in filename else ""
    if ext == "csv":
        return "csv"
    elif ext in ("xlsx", "xls"):
        return "xlsx"
    elif ext == "txt" and content:
        header = content[:500].decode("utf-8", errors="replace")
        if ":20:" in header:
            return "mt940"
        elif "camt.053" in header.lower() or "<BkToCstmrStmt>" in header:
            return "camt053"
        elif "1С" in header or "1C" in header or "Счет" in header:
            return "1c_export"
    return "csv"  # default


def parse_csv(file: BinaryIO) -> list[BankRow]:
    """Parse CSV bank export."""
    content = file.read().decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(content))
    rows = []
    for i, row in enumerate(reader):
        try:
            amount = abs(float(row.get("amount", row.get("sum", row.get("Сумма", 0)))))
            direction = "inflow" if float(row.get("amount", row.get("sum", 0))) > 0 else "outflow"
            rows.append(BankRow(
                external_id=row.get("id", row.get("external_id", str(uuid.uuid4()))),
                amount=amount,
                currency=row.get("currency", row.get("Валюта", "RUB")),
                direction=direction,
                transaction_date=_parse_date(row.get("date", row.get("Дата", ""))),
                value_date=_parse_date(row.get("value_date", "")),
                counterparty=row.get("counterparty", row.get("Контрагент")),
                account=row.get("account", row.get("Счет")),
                description=row.get("description", row.get("Назначение", row.get("Описание", ""))),
                purpose=row.get("purpose", row.get("Назначение")),
                reference=row.get("reference", row.get("Номер")),
                source_format="csv",
                raw=row,
            ))
        except (ValueError, TypeError) as e:
            continue  # skip malformed rows
    return rows


def parse_xlsx(file: BinaryIO) -> list[BankRow]:
    """Parse XLSX bank export (placeholder — requires openpyxl)."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        rows = []
        header = [str(c.value).lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(c is None for c in row):
                continue
            data = dict(zip(header, [str(v) if v else "" for v in row]))
            try:
                amount = abs(float(data.get("amount", data.get("sum", 0))))
                direction = "inflow" if float(data.get("amount", 0)) > 0 else "outflow"
                rows.append(BankRow(
                    external_id=data.get("id", str(uuid.uuid4())),
                    amount=amount,
                    currency=data.get("currency", "RUB"),
                    direction=direction,
                    transaction_date=_parse_date(data.get("date", "")),
                    counterparty=data.get("counterparty"),
                    description=data.get("description", data.get("purpose", "")),
                    source_format="xlsx",
                    raw=data,
                ))
            except (ValueError, TypeError):
                continue
        return rows
    except ImportError:
        raise ValueError("openpyxl required for XLSX parsing")


def parse_mt940(file: BinaryIO) -> list[BankRow]:
    """Parse MT940 (SWIFT) bank statement."""
    content = file.read().decode("utf-8", errors="replace")
    rows = []
    current_txn: dict = {}
    for line in content.splitlines():
        line = line.strip()
        if line.startswith(":61:"):
            # New transaction
            if current_txn:
                rows.append(_mt940_to_row(current_txn))
            # Parse :61: line
            parts = line[4:].split("}")
            raw = line[4:]
            value_date_str = raw[:6]
            entry_date_str = raw[6:10] if len(raw) > 6 else ""
            sign = raw[10] if len(raw) > 10 else "C"
            amount_str = raw[11:].split("N")[0] if "N" in raw[11:] else raw[11:25]
            try:
                amount = float(amount_str) / 100 if "." not in amount_str else float(amount_str)
            except ValueError:
                amount = 0
            current_txn = {
                "amount": amount,
                "direction": "inflow" if sign == "C" else "outflow",
                "value_date": _parse_swift_date(value_date_str),
                "entry_date": _parse_swift_date(entry_date_str) if len(entry_date_str) >= 6 else None,
                "raw": raw,
            }
        elif line.startswith(":86:") or (current_txn and line.startswith("~")):
            # Description
            desc = line[4:] if line.startswith(":86:") else line.replace("~", "")
            current_txn["description"] = current_txn.get("description", "") + " " + desc
    if current_txn:
        rows.append(_mt940_to_row(current_txn))
    return rows


def parse_camt053(file: BinaryIO) -> list[BankRow]:
    """Parse CAMT.053 XML bank statement."""
    try:
        import xml.etree.ElementTree as ET
    except ImportError:
        raise ValueError("xml.etree required for CAMT.053")
    content = file.read().decode("utf-8", errors="replace")
    rows = []
    try:
        root = ET.fromstring(content)
        ns = {"ns": "urn:iso:std:iso:20022:tech:xsd:camt.053.001.08"}
        for ntry in root.findall(".//ns:Ntry", ns):
            amt_str = ntry.findtext("ns:Amt", "", ns)
            cdt_dbt = ntry.findtext("ns:CdtDbtInd", "", ns)
            try:
                amount = abs(float(amt_str)) if amt_str else 0
            except ValueError:
                amount = 0
            rows.append(BankRow(
                external_id=ntry.findtext("ns:AcctSvcrRef", "", ns) or str(uuid.uuid4()),
                amount=amount,
                currency="RUB",
                direction="inflow" if cdt_dbt == "CRDT" else "outflow",
                transaction_date=date.today(),
                description=ntry.findtext(".//ns:Ustrd", "", ns),
                source_format="camt053",
                raw={"camt053": True},
            ))
    except ET.ParseError:
        raise ValueError("Invalid CAMT.053 XML")
    return rows


def parse_1c(file: BinaryIO) -> list[BankRow]:
    """Parse 1C bank export (tab-separated or CSV)."""
    content = file.read().decode("cp1251", errors="replace")
    rows = []
    for line in content.splitlines():
        if not line.strip() or line.startswith(";"):
            continue
        parts = line.split("\t")
        if len(parts) < 3:
            continue
        try:
            amount = abs(float(parts[5].replace(",", "."))) if len(parts) > 5 else 0
            rows.append(BankRow(
                external_id=str(uuid.uuid4()),
                amount=amount,
                currency="RUB",
                direction="inflow" if float(parts[5].replace(",", ".")) > 0 else "outflow" if len(parts) > 5 else "outflow",
                transaction_date=_parse_date(parts[1]) if len(parts) > 1 else date.today(),
                counterparty=parts[3] if len(parts) > 3 else None,
                description=parts[8] if len(parts) > 8 else parts[2] if len(parts) > 2 else "",
                source_format="1c_export",
                raw={"parts": parts},
            ))
        except (ValueError, IndexError):
            continue
    return rows


# ── Normalizer ─────────────────────────────────────────────────────────


def normalize(rows: list[BankRow]) -> list[BankRow]:
    """Normalize bank rows: dates, amounts, encoding, direction normalization."""
    for row in rows:
        # Normalize amount
        row.amount = round(abs(float(row.amount)), 2)
        # Normalize currency
        row.currency = (row.currency or "RUB").upper()[:3]
        # Normalize direction
        if row.direction in ("C", "CRDT", "credit", "приход", "поступление"):
            row.direction = "inflow"
        elif row.direction in ("D", "DBIT", "debit", "расход", "списание"):
            row.direction = "outflow"
        # Truncate descriptions
        if row.description:
            row.description = row.description.strip()[:500]
    return rows


# ── Fingerprint / Dedup ───────────────────────────────────────────────


def fingerprint(row: BankRow) -> str:
    """Deterministic fingerprint for dedup."""
    raw = f"{row.amount:.2f}|{row.direction}|{row.transaction_date}|{row.description or ''}|{row.counterparty or ''}|{row.external_id or ''}"
    return hashlib.sha256(raw.encode("utf-8", errors="replace")).hexdigest()


def dedup(rows: list[BankRow], existing_fingerprints: set[str] | None = None) -> tuple[list[BankRow], int]:
    """Deduplicate against existing fingerprints."""
    seen: set[str] = set(existing_fingerprints or [])
    unique: list[BankRow] = []
    dup_count = 0
    for row in rows:
        fp = fingerprint(row)
        if fp in seen:
            dup_count += 1
        else:
            seen.add(fp)
            unique.append(row)
    return unique, dup_count


# ── Preview ────────────────────────────────────────────────────────────


def preview(rows: list[BankRow], limit: int = 10) -> list[dict]:
    """Generate import preview (first N rows + summary)."""
    summary = {
        "total_rows": len(rows),
        "total_amount": sum(r.amount for r in rows),
        "inflow_count": sum(1 for r in rows if r.direction == "inflow"),
        "outflow_count": sum(1 for r in rows if r.direction == "outflow"),
        "date_range": (
            min(r.transaction_date for r in rows) if rows else None,
            max(r.transaction_date for r in rows) if rows else None,
        ),
    }
    preview_rows = [
        {
            "external_id": r.external_id,
            "amount": r.amount,
            "direction": r.direction,
            "date": str(r.transaction_date),
            "counterparty": r.counterparty,
            "description": r.description[:80] if r.description else "",
            "fingerprint": fingerprint(r)[:12],
        }
        for r in rows[:limit]
    ]
    return {"summary": summary, "rows": preview_rows}


# ── Import Orchestrator ───────────────────────────────────────────────


async def import_file(filename: str, content: bytes, company_id: str, confirm: bool = False) -> ImportResult:
    """Full import pipeline: detect → parse → normalize → dedup → preview → confirm.

    Args:
        filename: Original filename (for format detection)
        content: File bytes
        company_id: Target company
        confirm: If True, creates accounting_batch + accounting_events

    Returns:
        ImportResult with batch_id, stats, warnings
    """
    fmt = detect_format(filename, content)
    warnings: list[str] = []

    # Parse
    parser_func = {
        "csv": parse_csv,
        "xlsx": parse_xlsx,
        "mt940": parse_mt940,
        "camt053": parse_camt053,
        "1c_export": parse_1c,
    }.get(fmt)

    if not parser_func:
        raise ValueError(f"Unsupported format: {fmt}")

    rows = parser_func(io.BytesIO(content))
    if not rows:
        raise ValueError(f"No rows parsed from {filename}")

    # Normalize
    rows = normalize(rows)

    # Dedup
    pool = await get_pool()
    existing_fps: set[str] = set()
    async with pool.acquire() as conn:
        existing = await conn.fetch(
            "SELECT DISTINCT event_fingerprint FROM accounting.accounting_event WHERE company_id = $1",
            company_id,
        )
        for r in existing:
            if r["event_fingerprint"]:
                existing_fps.add(r["event_fingerprint"])

    unique, dups = dedup(rows, existing_fps)
    if dups > 0:
        warnings.append(f"{dups} duplicate rows skipped")

    # Preview
    prev = preview(unique)

    if not confirm:
        return ImportResult(
            batch_id="",
            events_created=0,
            duplicates=dups,
            warnings=warnings,
            preview=unique,
            fingerprint=hashlib.sha256(str(len(unique)).encode()).hexdigest()[:16],
        )

    # Confirm: create batch + events
    batch_id = str(uuid.uuid4())
    events_created = 0
    async with pool.acquire() as conn:
        await conn.execute(
                """INSERT INTO accounting.accounting_batch
                   (id, company_id, status, source, created_at)
                   VALUES ($1, $2, 'completed', $3, now())""",
                batch_id, company_id, f"bank_import:{fmt}",
            )

        for row in unique:
            fp = fingerprint(row)
            # Map direction to event_type
            event_type = EventType.SALE.value if row.direction == "inflow" else EventType.PURCHASE.value

            await conn.execute(
                """INSERT INTO accounting.accounting_event
                   (id, company_id, batch_id, event_type, event_date, amount,
                    currency, source_system, source_type, source_id,
                    description, processing_state, event_fingerprint,
                    is_current, version, created_at)
                   VALUES ($1, $2, $3, $4, $5, $6, $7,
                           'bank_import', $8, $9,
                           $10, 'new', $11,
                           true, 1, now())""",
                str(uuid.uuid4()), company_id, batch_id,
                event_type, row.transaction_date, row.amount, row.currency,
                fmt, row.external_id or "",
                row.description or "", fp,
            )
            events_created += 1

    return ImportResult(
        batch_id=batch_id,
        events_created=events_created,
        duplicates=dups,
        warnings=warnings,
        preview=unique,
    )


# ── Helpers ────────────────────────────────────────────────────────────


def _parse_date(val: str) -> date:
    """Parse various date formats."""
    if not val:
        return date.today()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y%m%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(val.strip(), fmt).date()
        except ValueError:
            continue
    return date.today()


def _parse_swift_date(val: str) -> date | None:
    """Parse SWIFT date (YYMMDD or YYYYMMDD)."""
    if len(val) == 6:
        try:
            return datetime.strptime(val, "%y%m%d").date()
        except ValueError:
            return None
    elif len(val) == 8:
        try:
            return datetime.strptime(val, "%Y%m%d").date()
        except ValueError:
            return None
    return None


def _mt940_to_row(txn: dict) -> BankRow:
    return BankRow(
        external_id=txn.get("raw", "")[:50] if txn.get("raw") else str(uuid.uuid4()),
        amount=txn.get("amount", 0),
        currency="RUB",
        direction=txn.get("direction", "outflow"),
        transaction_date=txn.get("entry_date") or txn.get("value_date") or date.today(),
        value_date=txn.get("value_date"),
        description=txn.get("description", "").strip()[:200],
        source_format="mt940",
        raw=txn,
    )

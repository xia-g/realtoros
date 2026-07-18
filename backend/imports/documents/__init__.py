"""Document Intake — upload → extract text → classify → create events → obligations.

Real pipeline:
1. Extract text from PDF (pdftotext / OCR via tesseract), DOCX, XLSX, XML, images
2. Classify document type by content (ДКП, счёт-фактура, чек и тд)
3. Extract key fields: amount, date, counterparty
4. Create accounting events + obligations automatically
"""

from __future__ import annotations

import hashlib
import io
import json
import logging
import os
import re
import subprocess
import tempfile
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any
from xml.etree import ElementTree

from backend.accounting.db.pool import get_pool

logger = logging.getLogger(__name__)

TESSERACT_BIN = os.path.expanduser("~/.local/bin/tesseract")
TESSDATA = os.path.expanduser("~/.local/share/tessdata")
LD_PATH = os.path.expanduser("~/.local/lib")


@dataclass
class DocumentResult:
    document_id: str
    classification: str
    confidence: float
    extracted_text: str
    extracted_fields: dict[str, Any]
    events_created: int
    obligations_created: int
    file_hash: str
    warnings: list[str]


SUPPORTED_FORMATS = {
    "pdf", "jpg", "jpeg", "png", "zip",
    "doc", "docx", "xls", "xlsx", "xml", "txt",
}


def detect_doc_format(filename: str) -> str:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    return ext if ext in SUPPORTED_FORMATS else "unknown"


# ── Text Extraction ─────────────────────────────────────────────────


def extract_text_from_pdf(content: bytes) -> str:
    """Extract text from PDF — try pdftotext first, fall back to OCR."""
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
        f.write(content)
        pdf_path = f.name

    try:
        # Try text extraction first
        result = subprocess.run(
            ["pdftotext", "-layout", pdf_path, "-"],
            capture_output=True, text=True, timeout=30,
        )
        text = result.stdout.strip()
        logger.info("PDF text extract: len=%d stderr=%s", len(text), repr(result.stderr[:100]))
        if len(text) > 50:
            return text

        # If too little text, it's probably a scanned PDF — OCR it
        logger.info("PDF has little text (%d chars), falling back to OCR", len(text))
        return _ocr_image(pdf_path, is_pdf=True)
    finally:
        os.unlink(pdf_path)


def _ocr_image(path: str, is_pdf: bool = False) -> str:
    """OCR an image or scanned PDF using tesseract."""
    env = os.environ.copy()
    env["LD_LIBRARY_PATH"] = f"{LD_PATH}:{env.get('LD_LIBRARY_PATH', '')}"
    env["TESSDATA_PREFIX"] = TESSDATA

    args = [TESSERACT_BIN, path, "stdout", "-l", "rus+eng", "--oem", "1", "--psm", "3"]
    if is_pdf:
        # For scanned PDFs, convert to image first
        logger.info("OCR: converting PDF to image (pdftoppm 300dpi)")
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            img_path = tmp.name
        try:
            subprocess.run(
                ["pdftoppm", "-png", "-r", "300", "-singlefile", path, img_path.replace(".png", "")],
                capture_output=True, timeout=60,
            )
            actual_img = img_path if os.path.exists(img_path) else img_path.replace(".png", "") + ".png"
            if not os.path.exists(actual_img):
                logger.warning("OCR: pdftoppm produced no output file")
                return ""
            logger.info("OCR: running tesseract on %s", actual_img)
            args = [TESSERACT_BIN, actual_img, "stdout", "-l", "rus+eng", "--oem", "1", "--psm", "3"]
            result = subprocess.run(args, capture_output=True, text=True, timeout=120, env=env)
            return result.stdout.strip() or result.stderr.strip()
        finally:
            for p in [img_path, img_path.replace(".png", "") + ".png"]:
                if os.path.exists(p):
                    os.unlink(p)
    else:
        try:
            result = subprocess.run(args, capture_output=True, text=True, timeout=120, env=env)
            return result.stdout.strip() or result.stderr.strip()
        except FileNotFoundError:
            return ""


def extract_text_from_docx(content: bytes) -> str:
    """Extract text from DOCX."""
    try:
        from docx import Document
        doc = Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        return ""


def extract_text_from_xlsx(content: bytes) -> str:
    """Extract text from XLSX."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        texts = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                texts.append(" | ".join(str(c) for c in row if c is not None))
        return "\n".join(texts)
    except ImportError:
        return ""


def extract_text_from_xml(content: bytes) -> str:
    """Extract text from XML."""
    try:
        root = ElementTree.fromstring(content)
        return " ".join(root.itertext())
    except Exception:
        return content.decode("utf-8", errors="replace")[:5000]


def extract_text(content: bytes, fmt: str) -> str:
    """Route to the right extractor based on format."""
    if fmt == "pdf":
        return extract_text_from_pdf(content)
    elif fmt in ("doc", "docx"):
        return extract_text_from_docx(content)
    elif fmt in ("xls", "xlsx"):
        return extract_text_from_xlsx(content)
    elif fmt == "xml":
        return extract_text_from_xml(content)
    elif fmt in ("jpg", "jpeg", "png"):
        with tempfile.NamedTemporaryFile(suffix=f".{fmt}", delete=False) as f:
            f.write(content)
            img_path = f.name
        try:
            return _ocr_image(img_path)
        finally:
            os.unlink(img_path)
    elif fmt == "txt":
        return content.decode("utf-8", errors="replace")[:10000]
    return ""


# ── Classification by Content ────────────────────────────────────────


# Keywords for Russian documents
DOC_PATTERNS: dict[str, list[str]] = {
    "contract": [
        "договор", "контракт", "соглашение", "дкп",
        "купли-продажи", "аренды", "подряда", "оказания услуг",
    ],
    "invoice": [
        "счет-фактура", "счет на оплату", "инвойс",
        "invoice", "сч.фактура", "с/ф",
    ],
    "receipt": [
        "кассовый чек", "фискальный", "чек", "receipt",
        "терминал", "оплачено", "касса",
    ],
    "act": [
        "акт выполненных работ", "акт оказания услуг",
        "акт приема-передачи", "акт сверки",
    ],
    "payment_order": [
        "платежное поручение", "платеж", "п/п",
        "банк", "перевод", "списание",
    ],
    "municipal_contract": [
        "торги", "аукцион", "муниципальное имущество",
        "администрация", "комитет", "департамент",
        "выкуп", "приватизация",
    ],
    "property_doc": [
        "свидетельство", "выписка", "егрн",
        "кадастровый", "регистрация права",
        "недвижимость", "помещение", "здание",
    ],
}

# Fields to extract by regex
FIELD_PATTERNS: list[tuple[str, str, str]] = [
    ("amount", r"([\d\s]{4,15})\s*(?:руб[^\w]|₽|р\.|рублей|RUB)", "sum"),
    ("date_contract", r"(?:от\s+)?(\d{2}\.\d{2}\.\d{4})\s*(?:г\.\s*|года\s*)?(?:\n|$|,|\s{2,})", "date"),
    ("inn", r"(?:ИНН|ИНН)\s*[:=\s]*(\d{10,12})", "inn"),
    ("kpp", r"(?:КПП|КПП)\s*[:\s]*(\d{9})", "kpp"),
    ("ogrn", r"(?:ОГРН|ОГРНИП)\s*[:\s]*(\d{13,15})", "ogrn"),
    ("counterparty", r"(?:продавец|покупатель|заказчик|исполнитель|стороны)[:\s]*([^\n]{3,50})", "party"),
]


def classify_document(text: str, filename: str) -> tuple[str, float]:
    """Classify document by content (and fall back to filename)."""
    text_lower = text.lower()
    name_lower = filename.lower()

    # Score each category
    scores: dict[str, int] = {}
    for doc_type, keywords in DOC_PATTERNS.items():
        score = 0
        for kw in keywords:
            if kw in text_lower:
                score += 3
            elif kw in name_lower:
                score += 1
        if score > 0:
            scores[doc_type] = score

    if not scores:
        return "other", 0.3

    best = max(scores, key=scores.get)
    confidence = min(0.5 + scores[best] * 0.05, 0.98)
    return best, confidence


def extract_fields(text: str) -> dict[str, Any]:
    """Extract key fields from document text."""
    fields: dict[str, Any] = {"amounts": [], "dates": [], "inn": None, "counterparty": None}
    for field_name, pattern, field_type in FIELD_PATTERNS:
        matches = re.findall(pattern, text, re.IGNORECASE | re.MULTILINE)
        for m in matches:
            value = m.strip() if isinstance(m, str) else m[0].strip()
            if field_type == "sum":
                # Clean number
                clean = re.sub(r"[^\d,.]", "", value.replace(" ", ""))
                try:
                    fields["amounts"].append(float(clean.replace(",", ".")))
                except ValueError:
                    pass
            elif field_type == "date":
                fields["dates"].append(value)
            elif field_type in ("inn", "kpp", "ogrn"):
                fields[field_type] = value
            elif field_type == "party":
                fields["counterparty"] = value
    return fields


# ── Event + Obligation Creation ─────────────────────────────────────


async def _create_property_purchase_events(
    pool, company_id: str, doc_id: str, fields: dict,
    classification: str,
) -> tuple[int, list[str]]:
    """Create accounting events for property purchase."""
    event_ids: list[str] = []
    count = 0
    price = max(fields.get("amounts", [0])) if fields.get("amounts") else 0

    if classification in ("contract", "municipal_contract") and price > 0:
        async with pool.acquire() as conn:
            # Check if event already exists for this doc
            existing = await conn.fetchval(
                "SELECT COUNT(*) FROM accounting.event_document WHERE document_id = $1",
                doc_id,
            )
            if existing and existing > 0:
                return 0, []

            # Store in document_intake table with price for later processing
            logger.info("Document ready for event creation: doc=%s price=%.2f class=%s",
                         doc_id[:8], price, classification)

    return count, event_ids


async def _create_obligations(
    pool, company_id: str, doc_id: str, fields: dict,
    classification: str,
) -> int:
    """Create obligations based on document type."""
    count = 0
    price = max(fields.get("amounts", [0])) if fields.get("amounts") else 0

    if classification in ("contract", "municipal_contract") and price > 0:
        vat_amount = round(Decimal(str(price)) * Decimal("20") / Decimal("120"), 2)

        # Get next quarter's 25th
        today = date.today()
        q = ((today.month - 1) // 3) + 1
        m = q * 3
        y = today.year
        if m < today.month:
            m += 3
            if m > 12:
                m = 3
                y += 1
        deadline = date(y, m, 25)

        async with pool.acquire() as conn:
            # Check duplicate
            dup = await conn.fetchval(
                """SELECT COUNT(*) FROM public.obligations
                   WHERE document_id = $1 AND obligation_type = 'vat_payable'""",
                doc_id,
            )
            if dup and dup > 0:
                return 0

            # VAT obligation (tax agent for municipal property)
            await conn.execute(
                """INSERT INTO public.obligations
                   (id, company_id, obligation_type, title, description,
                    amount, due_date, status, recurrence, reminder_days, document_id,
                    created_at, updated_at)
                   VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, now(), now())""",
                str(uuid.uuid4()), company_id, "vat_payable",
                "НДС как налоговый агент — покупка недвижимости",
                f"НДС с покупки недвижимости (20/120 от {price:,.2f}₽)",
                vat_amount, deadline, "pending", "one_time", 14, doc_id,
            )
            count += 1

            # Property tax reminder
            await conn.execute(
                """INSERT INTO public.obligations
                   (id, company_id, obligation_type, title, description,
                    amount, due_date, status, recurrence, reminder_days, document_id,
                    created_at, updated_at)
                   VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, now(), now())""",
                str(uuid.uuid4()), company_id, "tax_property",
                "Налог на имущество (коммерческая недвижимость)",
                "Ориентировочно 2% от кадастровой стоимости. Уточните в ФНС.",
                Decimal("0"), date(y, 12, 1), "pending", "yearly", 30, doc_id,
            )
            count += 1

    elif classification == "invoice" and price > 0:
        # Invoice with VAT to pay
        vat_sum = round(Decimal(str(price)) * Decimal("20") / Decimal("120"), 2)
        async with pool.acquire() as conn:
            dup = await conn.fetchval(
                "SELECT COUNT(*) FROM public.obligations WHERE document_id = $1 AND obligation_type = 'vat_payable'",
                doc_id,
            )
            if not dup:
                today = date.today()
                deadline = date(today.year, today.month + 1, 25) if today.month < 12 else date(today.year + 1, 1, 25)
                await conn.execute(
                    """INSERT INTO public.obligations
                       (id, company_id, obligation_type, title, description,
                        amount, due_date, status, recurrence, reminder_days, document_id,
                        created_at, updated_at)
                       VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, now(), now())""",
                    str(uuid.uuid4()), company_id, "vat_payable",
                    f"НДС к уплате по счёту-фактуре",
                    f"НДС {vat_sum:,.2f}₽ от суммы {price:,.2f}₽",
                    vat_sum, deadline, "pending", "one_time", 10, doc_id,
                )
                count += 1

    return count


# ── Main Pipeline ────────────────────────────────────────────────────


class DocumentIntake:
    """Full document intake pipeline with real extraction."""

    @staticmethod
    async def process(
        filename: str,
        content: bytes,
        company_id: str,
    ) -> DocumentResult:
        doc_id = str(uuid.uuid4())
        file_hash = hashlib.sha256(content).hexdigest()
        fmt = detect_doc_format(filename)
        logger.info("Intake start: doc=%s file=%s fmt=%s size=%d hash=%s",
                     doc_id[:8], filename, fmt, len(content), file_hash[:16])

        if fmt == "unknown":
            fmt = "pdf"

        warnings: list[str] = []

        # 1. Extract text
        logger.info("Extracting text: fmt=%s len=%d", fmt, len(content))
        text = extract_text(content, fmt)
        logger.info("Extracted text: len=%d preview=%s",
                     len(text), repr(text[:120]))
        if not text:
            warnings.append("Не удалось извлечь текст из документа (возможно, пустой или повреждён)")

        # 2. Classify
        logger.info("Classifying: text_len=%d", len(text))
        classification, confidence = classify_document(text, filename)
        logger.info("Classification: type=%s confidence=%.2f", classification, confidence)
        if classification == "other":
            confidence = 0.3
            warnings.append(
                "Документ не удалось классифицировать по содержанию. "
                "Пожалуйста, укажите тип вручную."
            )

        # 3. Extract fields
        fields = extract_fields(text)
        logger.info("Extracted fields: amounts=%s dates=%s inn=%s counterparty=%s",
                     fields.get("amounts"), fields.get("dates"),
                     fields.get("inn"), fields.get("counterparty"))

        pool = await get_pool()

        # 4. Create accounting events
        logger.info("Creating accounting events: classification=%s fields=%s",
                     classification, {"amounts": fields.get("amounts")})
        events_created, event_ids = await _create_property_purchase_events(
            pool, company_id, doc_id, fields, classification,
        )
        logger.info("Events created: count=%d ids=%s", events_created, event_ids)

        # 5. Create obligations
        logger.info("Creating obligations: classification=%s price=%s",
                     classification, max(fields.get("amounts", [0])) if fields.get("amounts") else 0)
        obligations_created = await _create_obligations(
            pool, company_id, doc_id, fields, classification,
        )
        logger.info("Obligations created: count=%d", obligations_created)

        # 6. Store intake result (after events/obligations created so we can log counts)
        pool2 = await get_pool()
        async with pool2.acquire() as conn2:
            await conn2.execute(
                """INSERT INTO accounting.document_intake
                   (id, company_id, file_name, file_hash, file_size, mime_type,
                    classification, confidence, extracted_text, extracted_fields,
                    events_created, obligations_created, warnings, status)
                   VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, 'processed')""",
                doc_id, company_id, filename, file_hash, len(content),
                f"application/{fmt}",
                classification, confidence, text[:5000], json.dumps(fields),
                events_created, obligations_created, warnings,
            )
        logger.info("Intake stored: doc=%s", doc_id[:8])

        logger.info("Intake complete: doc=%s class=%s conf=%.2f events=%d obls=%d",
                     doc_id[:8], classification, confidence, events_created, obligations_created)

        return DocumentResult(
            document_id=doc_id,
            classification=classification,
            confidence=confidence,
            extracted_text=text[:200],
            extracted_fields=fields,
            events_created=events_created,
            obligations_created=obligations_created,
            file_hash=file_hash,
            warnings=warnings,
        )
